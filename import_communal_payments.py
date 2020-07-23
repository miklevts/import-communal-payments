import csv
import dataclasses
import datetime
import io
import logging
from decimal import Decimal, DecimalException
from typing import List, Optional

from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction
from sentry_sdk import capture_exception

from apps.accounts.models import User
from apps.buildings.models import Apartment
from apps.payments import models as payment_models
from apps.payments.models import Currency
from apps.payments.services import PaymentsEmailService

log = logging.getLogger(__name__)


class ImportPaymentException(Exception):
    """Base exception for all ImportCommunalPaymentsService errors"""


class FileTypeError(ImportPaymentException):
    """Failed file type"""

    pass


class ParseFileError(ImportPaymentException):
    """Failed to parse file"""

    pass


@dataclasses.dataclass
class CommunalPayment:
    """
    Communal Payment dataclass for import
    """

    email: str
    payer: User = dataclasses.field(init=False)
    account_number: str
    apartment: Apartment = dataclasses.field(init=False)
    date: datetime.date
    ext_number: str = ""
    price: Decimal = Decimal("0")
    building: str = ""
    description: str = ""
    currency: payment_models.Currency = None

    def __post_init__(self):
        self.payer = self.get_payer_by_email()
        self.apartment = self.get_apartment_by_account()

    def save_payment(self):
        payment, created = payment_models.CommunalPayment.objects.update_or_create(
            payer=self.payer,
            apartment=self.apartment,
            date=self.date,
            ext_number=self.ext_number,
            defaults={"currency": self.currency, "price": self.price, "description": self.description},
        )
        return payment

    def get_payer_by_email(self):
        try:
            return User.objects.get(email=self.email)
        except User.DoesNotExist:
            raise ImportPaymentException(f"Payer by email = {self.email} not found")

    def get_apartment_by_account(self):
        try:
            return Apartment.objects.get(account_number=self.account_number)
        except Apartment.DoesNotExist:
            raise ImportPaymentException(f"Apartment by account number #{self.account_number} not found")


class ImportCommunalPaymentsService:
    """
    Import communal payments from file
    """

    cols = 8
    start_row = 2
    col_ext_number = 0
    col_month = 1
    col_building = 2
    col_apartment = 3
    col_account_number = 4
    col_email = 5
    col_description = 6
    col_price = 7

    def __init__(self, file):
        self.file = file
        self.payments: List[CommunalPayment] = []
        self.errors: List[ImportPaymentException] = []
        self.currency = Currency.objects.get(code=settings.S1_DEFAULT_COMMUNAL_PAYMENT_CURRENCY_CODE)

    def run(self):
        self.import_data_from_file()
        self.save_payments()
        return self.errors

    def save_payments(self):
        with transaction.atomic():
            created_payments = []
            for payment in self.payments:
                created_payment = payment.save_payment()
                created_payments.append(created_payment)
            for payment in created_payments:
                PaymentsEmailService.payer_new_communal_payment(payment)
                for lodger in payment.apartment.lodgers.all():
                    PaymentsEmailService.lodger_new_communal_payment(payment, lodger.resident)

    def import_data_from_file(self):
        """
        Reads a csv or xlsx file and compiles payments list
        """
        func = {"csv": self.get_data_from_cvs, "xlsx": self.get_data_from_xlsx}
        try:
            if isinstance(self.file, str):
                data = func[self.file.split(".")[-1].lower()]()
            elif isinstance(self.file, InMemoryUploadedFile):
                data = func[self.file.name.split(".")[-1].lower()]()
            else:
                ValueError("Error download file")
        except Exception as e:
            capture_exception(e)
            self.errors.append(FileTypeError(e))
            return

        for i, line in enumerate(data, 1):
            if i < self.start_row:
                continue
            try:
                payment = self.build_payment(line, i)
                self.payments.append(payment)
                log.debug(f"Line {i}: {payment}")
            except ImportPaymentException as e:
                self.errors.append(e)
                log.error(f"Line {i}: {e}")

    def get_data_from_cvs(self) -> List[List]:
        """
        Reads a csv file and returns a data list
        """
        if isinstance(self.file, InMemoryUploadedFile):
            decoded_file = self.file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            return list(csv.reader(io_string, delimiter=",", quotechar='"'))
        elif isinstance(self.file, str):
            with open(self.file, "r") as f:
                return list(csv.reader(f, delimiter=",", quotechar='"'))
        return []

    def get_data_from_xlsx(self) -> List[List]:
        """
        Reads a xlsx file and returns a data list
        """
        from openpyxl import load_workbook

        wb = load_workbook(self.file)
        sheet = wb.active
        if sheet is None:
            raise FileTypeError("File does not have sheet")

        lines = []
        row_num, is_row = 0, True
        while is_row:
            line = []
            row_num += 1
            for col_num in range(1, self.cols + 1):
                value = sheet.cell(row=row_num, column=col_num).value
                if col_num == 1 and not value:
                    is_row = False
                    break
                if is_row:
                    line.append(value)
            if len(line) > 0:
                lines.append(line)
        return lines

    def build_payment(self, line, i) -> Optional[CommunalPayment]:
        """
        Prepares data and returns an instance of `CommunalPayment`
        """
        if len(line) != self.cols:
            raise ParseFileError(f"Invalid number of columns: line={i} got={len(line)} expected={self.cols}")

        price_str = str(line[self.col_price]).strip().replace(",", ".")
        try:
            price = Decimal(price_str)
        except DecimalException:
            raise ParseFileError(f"Invalid price value format: got={price_str} expected='00.00'")

        date = line[self.col_month]
        if not isinstance(date, datetime.datetime):
            try:
                date = datetime.datetime.strptime(date.strip().lower(), "%d-%m-%Y")
            except ValueError:
                raise ParseFileError(f"Invalid month value format: got={date} expected='01-01-2001'")

        return CommunalPayment(
            email=str(line[self.col_email]).strip().lower(),
            account_number=str(line[self.col_account_number]).strip(),
            date=date,
            ext_number=str(line[self.col_ext_number]).strip().lower(),
            price=price,
            description=str(line[self.col_description]).strip().lower(),
            building=str(line[self.col_building]).strip().lower(),
            currency=self.currency,
        )
